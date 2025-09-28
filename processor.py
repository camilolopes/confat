
import io, re, unicodedata
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import pdfplumber

def _autosize(ws):
    for c in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2

def _write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=str(col))
    for i in range(len(df)):
        for j, col in enumerate(df.columns, start=start_col):
            val = df.iloc[i][col]
            ws.cell(row=start_row + 1 + i, column=j, value=(None if pd.isna(val) else val))

def _normalize_header(s):
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode("ascii")
    t = t.lower()
    t = re.sub(r"r\$|\(r\$?\)|currency|valor\s*\(.*?\)", "valor", t)
    t = re.sub(r"[^a-z0-9]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

def _coerce_brl(x):
    if pd.isna(x): return None
    s = str(x)
    s = s.replace("R$", "").replace(" ", "")
    s = re.sub(r"[^0-9,.-]", "", s)
    if s.count(",") == 1 and s.count(".") >= 1:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    elif s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    try:
        return float(s)
    except:
        try:
            return pd.to_numeric(s, errors="coerce")
        except:
            return None

def _build_pie_image_xl(series_df, title, text_fontsize=8, title_fontsize=11):
    total = series_df["Valor BRL"].sum()
    labels = [
        f"{cat}\n{val/total:.1%} • R$ {val:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        for cat, val in zip(series_df["Categoria"], series_df["Valor BRL"])
    ]
    plt.figure(figsize=(8, 8))
    plt.pie(series_df["Valor BRL"], labels=labels, startangle=90, colors=plt.cm.Set2.colors, textprops={"fontsize": text_fontsize})
    plt.title(title, fontsize=title_fontsize); plt.tight_layout()
    buf = io.BytesIO(); plt.savefig(buf, format="png", bbox_inches="tight"); plt.close(); buf.seek(0)
    pil_img = PILImage.open(buf); return XLImage(pil_img)

def _write_sheet_consol(wb, name, data, header_row=1):
    ws = wb.create_sheet(name)
    for j, col in enumerate(data.columns, start=1): ws.cell(row=header_row, column=j, value=str(col))
    for i in range(len(data)):
        for j, col in enumerate(data.columns, start=1):
            ws.cell(row=header_row + 1 + i, column=j, value=None if pd.isna(data.iloc[i][col]) else data.iloc[i][col])
    headers_idx = {ws.cell(row=header_row, column=i).value: i for i in range(1, data.shape[1] + 1)}
    if "Valor BRL" in headers_idx:
        c = headers_idx["Valor BRL"]
        for r in range(header_row + 1, header_row + 1 + len(data)): ws.cell(row=r, column=c).number_format = u'R$ #,##0.00'
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"; _autosize(ws); return ws

def _add_index_and_order(wb, cards_display):
    if "Índice" in wb.sheetnames: del wb["Índice"]
    ws_idx = wb.create_sheet("Índice", 0); ws_idx["A1"] = "📑 Índice de Navegação"; row = 3
    for name in ["Consolidado Cartão","Consolidado Estabelecimento","Consolidado Cat por Cartão","Resumo Fatura","Devoluções","Parcelas Ativas"]:
        if name in wb.sheetnames:
            ws_idx.cell(row=row, column=1, value=name); ws_idx.cell(row=row, column=1).hyperlink = f"#'{name}'!A1"; ws_idx.cell(row=row, column=1).style = "Hyperlink"; row += 1
    ws_idx.cell(row=row, column=1, value="---"); row += 1
    ws_idx.cell(row=row, column=1, value="Cartões (Mapa de Calor – Top 3 + Outras):"); row += 1
    for display, sheet_name in cards_display:
        ws_idx.cell(row=row, column=1, value=display); ws_idx.cell(row=row, column=1).hyperlink = f"#'{sheet_name}'!A1"; ws_idx.cell(row=row, column=1).style = "Hyperlink"; row += 1
    ws_idx.column_dimensions["A"].width = 65
    desired_after_index = ["Consolidado Cartão","Consolidado Estabelecimento","Consolidado Cat por Cartão"]
    current = wb.sheetnames; ordered = ["Índice"] + [s for s in desired_after_index if s in current]
    for s in current:
        if s not in ordered: ordered.append(s)
    wb._sheets = [wb[s] for s in ordered]

# --------- C6 (Excel) ---------
def _pick_sheet_and_dataframe_c6(file_bytes):
    bio = io.BytesIO(file_bytes)
    try:
        df = pd.read_excel(bio, sheet_name="Transações Originais", header=0)
        if df is not None and not df.empty: return df
    except Exception: pass
    bio.seek(0); xl = pd.ExcelFile(bio); best_sheet = xl.sheet_names[0]; best_score = -1
    for sh in xl.sheet_names:
        try: df5 = xl.parse(sh, header=0, nrows=5)
        except Exception: continue
        norm_cols = [_normalize_header(c) for c in df5.columns]; score = 0
        for tokens in [{"nome","cartao"},{"final","cartao"},{"categoria"},{"descricao"},{"valor"}]:
            if any(all(tok in h for tok in tokens) for h in norm_cols): score += 1
        if score > best_score: best_score = score; best_sheet = sh
    bio.seek(0); df = pd.read_excel(bio, sheet_name=best_sheet, header=0)
    first_rows = min(8, len(df))
    for r in range(first_rows):
        row_vals = df.iloc[r].tolist(); norm = [_normalize_header(v) for v in row_vals]
        conds = [any("nome" in h and "cartao" in h for h in norm), any("final" in h and "cartao" in h for h in norm),
                 any("categoria" in h for h in norm), any("descricao" in h or "estabelecimento" in h for h in norm),
                 any("valor" in h for h in norm)]
        if sum(conds) >= 3:
            new_cols = [str(v) for v in row_vals]; df = df.iloc[r+1:].reset_index(drop=True); df.columns = new_cols; break
    return df

def build_processed_workbook_c6(file_bytes: bytes) -> bytes:
    df = _pick_sheet_and_dataframe_c6(file_bytes)
    norm_map = {_normalize_header(c): c for c in df.columns}
    def find_col(*tokens_sets):
        for norm, orig in norm_map.items():
            for tokens in tokens_sets:
                if all(tok in norm for tok in tokens): return orig
        return None
    col_nome = find_col({"nome","cartao"}, {"portador"}, {"titular"})
    col_final = find_col({"final","cartao"}, {"final","****"}, {"cartao","final"})
    col_categoria = find_col({"categoria"})
    col_descricao = find_col({"descricao"}, {"estabelecimento"}, {"loja"}, {"merchant"})
    col_valor = find_col({"valor"})
    col_data = find_col({"data"})
    required = {"Nome no Cartão": col_nome, "Final do Cartão": col_final, "Categoria": col_categoria, "Descrição": col_descricao, "Valor BRL": col_valor}
    missing = [k for k,v in required.items() if v is None]
    if missing: raise ValueError(f"Não encontrei colunas: {missing}. Títulos encontrados: {list(df.columns)}")
    df = df.rename(columns={col_nome:"Nome no Cartão", col_final:"Final do Cartão", col_categoria:"Categoria", col_descricao:"Descrição", col_valor:"Valor BRL", **({col_data:"Data"} if col_data else {})})
    df["Valor BRL"] = df["Valor BRL"].apply(_coerce_brl)
    if "Data" in df.columns: df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    def last4(x):
        s = str(x); m = re.findall(r"(\d{4})", s)
        return m[-1] if m else s[-4:]
    df["Final do Cartão"] = df["Final do Cartão"].apply(last4)
    # Extract parcela from descrição and enrich
    if "Descrição" in df.columns:
        parc, clean_desc = [], []
        for d in df["Descrição"]:
            desc = None if pd.isna(d) else str(d)
            if desc is None: clean_desc.append(d); parc.append(None); continue
            mm = re.search(r"(?:Parc(?:ela)?\s*)?(\d{1,2}\s*/\s*\d{1,2})\s*$", desc, flags=re.IGNORECASE)
            if mm:
                parc.append(mm.group(1))
                clean_desc.append(re.sub(r"(?:Parc(?:ela)?\s*)?(\d{1,2}\s*/\s*\d{1,2})\s*$", "", desc, flags=re.IGNORECASE).rstrip(" -–,"))
            else:
                parc.append(None); clean_desc.append(desc)
        df["Parcela"] = parc; df["Descrição"] = clean_desc
    df = _enrich_parcelamento_columns(df)
    return _build_excel_from_transactions(df)

# --------- Nubank (PDF) ---------
def _clean_person_name_candidate(s):
    if not s: return None
    t = str(s).strip(); t = re.sub(r"[^\w\s\.\-Á-Üá-ü]", " ", t, flags=re.UNICODE); t = re.sub(r"\s+", " ", t).strip()
    low = t.lower()
    blacklist = ["olá","ola","bem vindo","bem-vindo","sua fatura","resumo","nubank","cartao","cartão","fatura","limite","vencimento","valor","pagamento","pdf","visa","mastercard","credito","crédito","debito","débito","titular:","nome:","endereco","endereço"]
    if any(k in low for k in blacklist): return None
    if any(ch.isdigit() for ch in t): return None
    words = [w for w in re.split(r"[\s\.]+", t) if w]
    if not (2 <= len(words) <= 6): return None
    if any(len(w) < 2 for w in words): return None
    letters_ratio = sum(c.isalpha() for c in t) / max(1, len(t))
    if letters_ratio < 0.7: return None
    lowers = {"da","de","do","dos","das","e"}; fixed = []
    for i,w in enumerate(words):
        wl = w.lower()
        if i>0 and wl in lowers: fixed.append(wl)
        else: fixed.append(wl.capitalize())
    name = " ".join(fixed)
    if " " not in name: return None
    if any(len(tok) < 2 for tok in name.split()): return None
    return name

def _guess_holder_from_header(full_text):
    m = re.search(r"(?:Titular|Nome)\s*[:\-]\s*([A-Za-zÁ-Üá-ü\.\s]+)", full_text)
    if m:
        cand = _clean_person_name_candidate(m.group(1))
        if cand: return cand
    for line in full_text.splitlines()[:80]:
        cand = _clean_person_name_candidate(line)
        if cand: return cand
    return "Nubank"

def _extract_holder_candidates_from_pages(file_bytes: bytes):
    bio = io.BytesIO(file_bytes); cands = []
    with pdfplumber.open(bio) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            lines = [l.strip() for l in txt.splitlines()[:12] if l and l.strip()]
            for t in lines:
                s = re.sub(r"[^\w\s\.\-Á-Üá-ü]", " ", t, flags=re.UNICODE).strip()
                if len(s) < 8: continue
                low = s.lower()
                if any(k in low for k in ["olá","ola","nubank","fatura","cartao","cartão","resumo","vencimento","pagamento","limite","valor"]): continue
                letters_total = sum(ch.isalpha() for ch in s)
                if letters_total == 0: continue
                upper_ratio = sum(ch.isupper() for ch in s if ch.isalpha()) / letters_total
                if upper_ratio < 0.8: continue
                if len(s.split()) < 2: continue
                words = re.split(r"\s+", s); lowers = {"da","de","do","dos","das","e"}; fixed = []
                for i,w in enumerate(words):
                    wl = w.lower()
                    if i>0 and wl in lowers: fixed.append(wl)
                    else: fixed.append(wl.capitalize())
                name = " ".join(fixed); cands.append(name)
    return cands

def _pt_month_to_num(m):
    m = (m or "").strip().lower()
    mapa = {"jan":1,"janeiro":1,"fev":2,"fevereiro":2,"mar":3,"marco":3,"março":3,"abr":4,"abril":4,"mai":5,"maio":5,"jun":6,"junho":6,"jul":7,"julho":7,"ago":8,"agosto":8,"set":9,"setembro":9,"sep":9,"out":10,"outubro":10,"nov":11,"novembro":11,"dez":12,"dezembro":12}
    return mapa.get(m)

def _parse_pt_date_token(tok, ref_year=None):
    tok = str(tok).strip()
    if not tok: return None
    m = re.match(r"^(\d{1,2})\s+([A-Za-zÀ-Üà-ü]{3,})(?:\s+(\d{4}))?$", tok, flags=re.IGNORECASE)
    if m:
        d = int(m.group(1)); mon = _pt_month_to_num(m.group(2)); y = int(m.group(3)) if m.group(3) else (ref_year or pd.Timestamp.today().year)
        if mon:
            try: return pd.Timestamp(year=y, month=mon, day=d)
            except Exception: return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", tok)
    if m:
        d = int(m.group(1)); mon = int(m.group(2)); y = m.group(3)
        if y is None: y = ref_year or pd.Timestamp.today().year
        else: y = int(y);  y = y + 2000 if y < 100 else y
        try: return pd.Timestamp(year=y, month=mon, day=d)
        except Exception: return None
    return None

def _extract_parcela(desc):
    if desc is None: return None, None
    s = str(desc)
    m = re.search(r"(?:Parcela\s*)?(\d{1,2}/\d{1,2})\s*$", s, flags=re.IGNORECASE)
    if m:
        parcela = m.group(1); s = s[:m.start()].rstrip(" -–,"); return s, parcela
    return s, None

def _parse_parcela_fields(parcela_str):
    if parcela_str is None: return (None, None)
    s = str(parcela_str).strip().lower(); s = re.sub(r"parc(?:ela)?\s*", "", s); s = s.replace(" de ", "/")
    m = re.search(r"(\d{1,2})\s*/\s*(\d{1,2})", s)
    if not m: return (None, None)
    try: return int(m.group(1)), int(m.group(2))
    except: return (None, None)

def _categorize(desc):
    if not desc: return "Outros"
    s = desc.lower()
    rules = [("porto seguro","Seguro"),("seguro","Seguro"),("ifood","Alimentação"),("pizza","Alimentação"),("padaria","Alimentação"),("rest","Alimentação"),
             ("uber","Transporte"),(" 99","Transporte"),("cabify","Transporte"),
             ("posto","Combustível"),("ipiranga","Combustível"),("shell","Combustível"),("br mania","Combustível"),
             ("mercadolivre","Marketplace"),("amazon","Marketplace"),("magalu","Marketplace"),("submarino","Marketplace"),("americanas","Marketplace"),
             ("netflix","Assinaturas"),("spotify","Assinaturas"),("youtube","Assinaturas"),
             ("tim","Telefonia"),("vivo","Telefonia"),("claro","Telefonia"),("oi ","Telefonia"),
             ("drog","Saúde"),("farm","Saúde"),("laborat","Saúde"),
             ("enel","Utilities"),("cpfl","Utilities"),("sabesp","Utilities"),("energia","Utilities"),("light","Utilities"),
             ("academ","Fitness"),("academia","Fitness")]
    for k,v in rules:
        if k in s: return v
    return "Outros"

def _enrich_parcelamento_columns(df):
    if 'Parcela' not in df.columns: df['Parcela'] = None
    atual_list, total_list = [], []
    for val in df['Parcela']:
        n_atual, n_total = _parse_parcela_fields(val); atual_list.append(n_atual); total_list.append(n_total)
    df['Parcela Nº'] = atual_list; df['Qtde Parcelas'] = total_list
    def rest(a,t): return None if (a is None or t is None) else max(t-a,0)
    df['Restantes'] = [rest(a,t) for a,t in zip(df['Parcela Nº'], df['Qtde Parcelas'])]
    def ult(a,t): return None if (a is None or t is None) else ("Sim" if a>=t else "Não")
    df['É Última?'] = [ult(a,t) for a,t in zip(df['Parcela Nº'], df['Qtde Parcelas'])]
    if 'Data' in df.columns: base_dates = pd.to_datetime(df['Data'], errors='coerce')
    else: base_dates = pd.Series([pd.NaT]*len(df))
    term_list = []
    for dt, rem in zip(base_dates, df['Restantes']):
        if pd.isna(dt) or rem is None: term_list.append(None)
        else:
            try: term_list.append((dt + pd.DateOffset(months=int(rem))).date())
            except Exception: term_list.append(None)
    df['Término Estimado'] = term_list; return df

def _parse_nubank_pdf(file_bytes: bytes) -> pd.DataFrame:
    bio = io.BytesIO(file_bytes); texts = []
    with pdfplumber.open(bio) as pdf:
        for page in pdf.pages: texts.append(page.extract_text() or "")
    full = "\n".join(texts)
    cands = _extract_holder_candidates_from_pages(file_bytes)
    if cands:
        from collections import Counter; holder = Counter(cands).most_common(1)[0][0]
    else:
        holder = _guess_holder_from_header(full)
    last4 = "0000"; m_last4 = re.search(r"•{2,}\s*(\d{4})", full)
    if not m_last4: m_last4 = re.search(r"(\d{4})\s*(?:•|\*{2,}|x{2,})?\s*$", full, flags=re.MULTILINE)
    if m_last4: last4 = m_last4.group(1)
    current_year = pd.Timestamp.today().year; rows = []
    for line in full.splitlines():
        line = line.strip()
        if not line: continue
        m = re.match(r"^(\d{1,2}\s+[A-Za-zÀ-Üà-ü]{3,}|\d{1,2}/\d{1,2}(?:/\d{2,4})?)\s+(.*)$", line)
        if not m: continue
        date_tok = m.group(1); rest = m.group(2)
        m_val = re.search(r"([\-–]?\s*R?\$?\s*[\d\.\,]+)\s*$", rest)
        if not m_val: continue
        val_str = m_val.group(1); left = rest[: m_val.start()].strip()
        desc_clean, parcela = _extract_parcela(left)
        neg_kw = r"(pagamento|estorno|ajuste|cr[eé]dito)"
        sign = -1 if (val_str.strip().startswith("-") or re.search(neg_kw, (desc_clean or ""), re.I)) else 1
        try: valor = sign * float(val_str.replace("R$", "").replace("–", "-").replace(".", "").replace(",", ".").strip())
        except: continue
        dt = _parse_pt_date_token(date_tok, ref_year=current_year)
        rows.append({"Data": dt,"Nome no Cartão": holder,"Final do Cartão": last4,"Categoria": _categorize(desc_clean),"Descrição": desc_clean,"Parcela": parcela,"Valor BRL": valor})
    df = pd.DataFrame(rows)
    if df.empty:
        bio.seek(0)
        with pdfplumber.open(bio) as pdf:
            for page in pdf.pages:
                tbl = page.extract_table()
                if not tbl: continue
                header = [str(x) for x in tbl[0]]
                for row in tbl[1:]:
                    row = [None if x is None else str(x) for x in row]
                    data = row[0] if len(row)>0 else None
                    descricao = row[1] if len(row)>1 else None
                    valor = row[-1] if len(row)>0 else None
                    if not (data and descricao and valor): continue
                    try: v = float(valor.replace("R$", "").replace(".", "").replace(",", "."))
                    except: continue
                    desc_clean, parcela = _extract_parcela(descricao)
                    df = pd.concat([df, pd.DataFrame([{"Data": _parse_pt_date_token(data, ref_year=current_year),"Nome no Cartão": holder, "Final do Cartão": last4,"Categoria": _categorize(desc_clean),"Descrição": desc_clean, "Parcela": parcela, "Valor BRL": v}])], ignore_index=True)
        if not df.empty:
            try: df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)
            except Exception: pass
    if "Valor BRL" in df.columns: df["Valor BRL"] = pd.to_numeric(df["Valor BRL"], errors="coerce")
    df = _enrich_parcelamento_columns(df); return df

def build_processed_workbook_nubank(file_bytes: bytes) -> bytes:
    df = _parse_nubank_pdf(file_bytes)
    return _build_excel_from_transactions(df)

# --------- Workbook builder ---------
def _build_excel_from_transactions(df: pd.DataFrame) -> bytes:
    df_pos = df[df["Valor BRL"] > 0].copy(); df_neg = df[df["Valor BRL"] < 0].copy()
    consol_cartao = (df_pos.groupby(["Final do Cartão","Nome no Cartão","Descrição"], as_index=False)["Valor BRL"].sum().sort_values(["Final do Cartão","Valor BRL"], ascending=[True, False]).rename(columns={"Nome no Cartão":"Nome do Portador"}))
    consol_estab = (df_pos.groupby(["Nome no Cartão","Final do Cartão","Descrição"], as_index=False)["Valor BRL"].sum().sort_values(["Nome no Cartão","Final do Cartão","Valor BRL"], ascending=[True, True, False]).rename(columns={"Nome no Cartão":"Nome do Portador"}))
    consol_cat_cartao = (df_pos.groupby(["Final do Cartão","Nome no Cartão","Categoria"], as_index=False)["Valor BRL"].sum().sort_values(["Final do Cartão","Valor BRL"], ascending=[True, False]).rename(columns={"Nome no Cartão":"Nome do Portador"}))
    resumo = pd.DataFrame({"Total Fatura (R$)":[df["Valor BRL"].sum()],"Total Sem Devoluções (R$)":[df_pos["Valor BRL"].sum()],"Total Devoluções (R$)":[df_neg["Valor BRL"].sum()]})

    wb = Workbook(); default = wb.active; wb.remove(default)
    _write_sheet_consol(wb, "Consolidado Cartão", consol_cartao)
    ws_ce = _write_sheet_consol(wb, "Consolidado Estabelecimento", consol_estab, header_row=2)
    ws_ce.insert_rows(1); ws_ce["A1"] = "NOTA: 'Final do Cartão' = últimos 4 dígitos; 'Nome do Portador' = nome impresso. Somente valores positivos."; ws_ce.freeze_panes = "A3"
    _write_sheet_consol(wb, "Consolidado Cat por Cartão", consol_cat_cartao)

    ws_dev = wb.create_sheet("Devoluções")
    cols_dev = ["Data","Nome no Cartão","Final do Cartão","Categoria","Descrição","Parcela","Valor BRL"]
    present = [c for c in cols_dev if c in df.columns]
    for j, col in enumerate(present, start=1): ws_dev.cell(row=1, column=j, value=str(col if col != "Nome no Cartão" else "Nome do Portador"))
    for i in range(len(df_neg)):
        for j, col in enumerate(present, start=1):
            val = df_neg.iloc[i][col]; ws_dev.cell(row=2+i, column=j, value=None if pd.isna(val) else val)
    if "Valor BRL" in present:
        col_idx = present.index("Valor BRL") + 1
        for r in range(2, ws_dev.max_row + 1): ws_dev.cell(row=r, column=col_idx).number_format = u'R$ #,##0.00'
    ws_dev.auto_filter.ref = f"A1:{get_column_letter(ws_dev.max_column)}{ws_dev.max_row}"; _autosize(ws_dev)

    ws_rf = wb.create_sheet("Resumo Fatura")
    ws_rf.cell(row=1, column=1, value="Total Fatura (R$)"); ws_rf.cell(row=1, column=2, value=resumo.iloc[0,0])
    ws_rf.cell(row=2, column=1, value="Total Sem Devoluções (R$)"); ws_rf.cell(row=2, column=2, value=resumo.iloc[0,1])
    ws_rf.cell(row=3, column=1, value="Total Devoluções (R$)"); ws_rf.cell(row=3, column=2, value=resumo.iloc[0,2])
    for r in range(1,4): ws_rf.cell(row=r, column=2).number_format = u'R$ #,##0.00'; _autosize(ws_rf)

    ws_to = wb.create_sheet("Transações Originais"); _write_df(ws_to, df); ws_to.sheet_state = "hidden"

    df_pos = df[df["Valor BRL"] > 0].copy()
    holder_map = (df_pos.groupby(["Final do Cartão","Nome no Cartão"])["Valor BRL"].sum().reset_index().sort_values(["Final do Cartão","Valor BRL"], ascending=[True, False]).drop_duplicates(subset=["Final do Cartão"]).set_index("Final do Cartão")["Nome no Cartão"].to_dict())
    cats_por_cartao = df_pos.groupby("Final do Cartão")["Categoria"].nunique().to_dict()
    gastos_por_cartao_cat = (df_pos.groupby(["Final do Cartão","Categoria"], as_index=False)["Valor BRL"].sum())

    cards_display = []
    for final_cartao, grupo in gastos_por_cartao_cat.groupby("Final do Cartão"):
        if grupo.shape[0] == 0: continue
        sheet_name = f"Cartão {final_cartao}"; ws_card = wb.create_sheet(sheet_name)
        ws_card["A1"] = f"Mapa de Calor - Cartão {final_cartao} (Top 3 + Outras)"
        tabela = grupo.sort_values("Valor BRL", ascending=False).reset_index(drop=True)
        if tabela.shape[0] > 3:
            top3 = tabela.head(3).copy(); outras_val = float(tabela["Valor BRL"].sum() - top3["Valor BRL"].sum())
            if outras_val > 0: top3 = pd.concat([top3, pd.DataFrame([{"Categoria":"Outras","Valor BRL":outras_val}])], ignore_index=True)
            tabela = top3
        holder = holder_map.get(str(final_cartao), "")
        chart_title = f"Distribuição de Gastos – Cartão {final_cartao}";  chart_title += f" – {holder}" if holder else ""
        img = _build_pie_image_xl(tabela, chart_title, text_fontsize=8, title_fontsize=11); ws_card.add_image(img, "A3")
        if cats_por_cartao.get(final_cartao, 0) <= 2: ws_card.sheet_state = "hidden"
        display_name = sheet_name + (f" – {holder}" if holder else ""); cards_display.append((display_name, sheet_name))

    # Parcelas Ativas
    df_parc = df.copy()
    if set(["Valor BRL","Parcela Nº","Qtde Parcelas","Restantes"]).issubset(df_parc.columns):
        mask_active = (df_parc["Valor BRL"] > 0) & df_parc["Parcela Nº"].notna() & df_parc["Qtde Parcelas"].notna() & (df_parc["Restantes"].fillna(0) > 0)
        parc_active = df_parc[mask_active].copy()
        if not parc_active.empty:
            parc_active["Compromisso Futuro (R$)"] = parc_active["Valor BRL"] * parc_active["Restantes"]
            cols = [c for c in ["Nome no Cartão","Final do Cartão","Descrição","Parcela","Parcela Nº","Qtde Parcelas","Restantes","Valor BRL","Compromisso Futuro (R$)","Término Estimado","Data","Categoria"] if c in parc_active.columns]
            ws_pa = wb.create_sheet("Parcelas Ativas")
            for j, col in enumerate(cols, start=1): ws_pa.cell(row=1, column=j, value=str(col if col != "Nome no Cartão" else "Nome do Portador"))
            for i in range(len(parc_active)):
                for j, col in enumerate(cols, start=1): ws_pa.cell(row=2+i, column=j, value=None if pd.isna(parc_active.iloc[i][col]) else parc_active.iloc[i][col])
            for hdr in ["Valor BRL","Compromisso Futuro (R$)"]:
                if hdr in cols:
                    cidx = cols.index(hdr)+1
                    for r in range(2, ws_pa.max_row+1): ws_pa.cell(row=r, column=cidx).number_format = u'R$ #,##0.00'
            ws_pa.auto_filter.ref = f"A1:{get_column_letter(ws_pa.max_column)}{ws_pa.max_row}"; _autosize(ws_pa)
            total_future = float(parc_active["Compromisso Futuro (R$)"].sum())
            ws_rf.cell(row=5, column=1, value="Total Compromissos Futuros (Parcelas)"); ws_rf.cell(row=5, column=2, value=total_future); ws_rf.cell(row=5, column=2).number_format = u'R$ #,##0.00'
            try:
                brk = parc_active.groupby(["Final do Cartão","Nome no Cartão"])["Compromisso Futuro (R$)"].sum().reset_index()
                ws_rf.cell(row=7, column=1, value="Compromissos por Cartão (final / portador)"); rr = 8
                for _, row in brk.iterrows():
                    ws_rf.cell(row=rr, column=1, value=f"{row['Final do Cartão']} – {row['Nome no Cartão']}")
                    ws_rf.cell(row=rr, column=2, value=float(row["Compromisso Futuro (R$)"])); ws_rf.cell(row=rr, column=2).number_format = u'R$ #,##0.00'
                    rr += 1
            except Exception: pass
            _autosize(ws_rf)

    _add_index_and_order(wb, cards_display)

    out_io = io.BytesIO(); wb.save(out_io); out_io.seek(0); return out_io.getvalue()
